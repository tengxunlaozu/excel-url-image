"""
Excel 图片处理器 - PySide6 GUI 版
将 Excel 中指定列的图片链接下载并插入为图片
"""

import sys
import os
import re
import requests
import tempfile
from datetime import datetime
from urllib.parse import urlparse, unquote

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QLineEdit, QPushButton, QProgressBar,
    QTextEdit, QFileDialog, QMessageBox, QFrame, QSizePolicy
)
from PySide6.QtCore import Qt, QThread, Signal, QSize
from PySide6.QtGui import QFont, QIcon, QColor, QPalette, QTextCursor

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D


# ============================================================
# 后台处理线程
# ============================================================
class WorkerThread(QThread):
    progress = Signal(int, int)       # 当前, 总数
    log = Signal(str)                 # 日志消息
    finished = Signal(bool, str)      # 成功/失败, 消息

    def __init__(self, input_path, output_path, col_letter, parent=None):
        super().__init__(parent)
        self.input_path = input_path
        self.output_path = output_path
        self.col_letter = col_letter.strip().upper()
        self.img_dir = os.path.join(os.getcwd(), "excel_images")

    def run(self):
        try:
            self._process()
        except Exception as e:
            self.finished.emit(False, f"处理异常: {e}")

    def _process(self):
        # 确保图片目录存在
        os.makedirs(self.img_dir, exist_ok=True)
        self.log.emit(f"📂 图片保存目录: {self.img_dir}")

        # 列字母 → 列索引 (1-based)
        col_idx = 0
        for ch in self.col_letter:
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
        self.log.emit(f"📌 图片链接所在列: {self.col_letter}列 (第{col_idx}列)")

        # 加载工作簿
        self.log.emit(f"📂 加载文件: {self.input_path}")
        wb = load_workbook(self.input_path)
        ws = wb.active

        max_row = ws.max_row
        if max_row < 2:
            self.finished.emit(False, "没有数据行")
            return

        # 设置行高
        row_height = 100
        for r in range(2, max_row + 1):
            ws.row_dimensions[r].height = row_height

        # 设置列宽
        ws.column_dimensions[self.col_letter].width = 25

        # 收集所有需要处理的行
        tasks = []
        for row in range(2, max_row + 1):
            cell_val = ws.cell(row=row, column=col_idx).value
            urls = self._extract_urls(cell_val)
            if urls:
                tasks.append((row, urls))

        if not tasks:
            self.finished.emit(False, "没有找到图片链接")
            return

        total = len(tasks)
        total_images = 0
        failed_downloads = 0
        self.log.emit(f"🔍 共找到 {total} 行需要处理")

        # 处理每一行
        for idx, (row, urls) in enumerate(tasks):
            self.log.emit(f"\n{'─'*40}")
            self.log.emit(f"📌 第 {row} 行，发现 {len(urls)} 张图片链接")
            self.progress.emit(idx + 1, total)

            tmp_files = []
            for img_idx, url in enumerate(urls):
                self.log.emit(f"  ⬇ 下载: {url}")
                tmp_path, saved_name = self._download_image(url, row, img_idx)
                if tmp_path:
                    tmp_files.append((tmp_path, saved_name))
                    total_images += 1
                else:
                    failed_downloads += 1

            if not tmp_files:
                continue

            # 计算图片尺寸
            img_sizes = []
            total_width_px = 0
            margin = 5
            img_height = 80

            for tmp_path, _ in tmp_files:
                xl_img = XLImage(tmp_path)
                scale = img_height / xl_img.height
                w = int(xl_img.width * scale)
                img_sizes.append((w, img_height, xl_img))
                total_width_px += w + margin
            total_width_px -= margin

            # 调整列宽
            needed_col = max(15, total_width_px / 7 + 2)
            ws.column_dimensions[self.col_letter].width = needed_col

            # 插入图片
            emu_per_px = 9525
            for img_idx, (tmp_path, saved_name) in enumerate(tmp_files):
                w, h, xl_img = img_sizes[img_idx]
                xl_img.width = w
                xl_img.height = h

                offset_x = 0
                for i in range(img_idx):
                    offset_x += (img_sizes[i][0] + margin) * emu_per_px

                offset_y = int((row_height - img_height) / 2 * emu_per_px)

                anchor = OneCellAnchor(
                    _from=AnchorMarker(
                        col=col_idx - 1,
                        row=row - 1,
                        colOff=offset_x,
                        rowOff=offset_y
                    ),
                    ext=XDRPositiveSize2D(
                        cx=int(w * emu_per_px),
                        cy=int(h * emu_per_px)
                    )
                )
                xl_img.anchor = anchor
                ws.add_image(xl_img)
                self.log.emit(f"  ✅ 插入图片: {saved_name} ({w}x{h}px)")

            # 更新进度
            self.progress.emit(idx + 1, total)

        # 保存
        wb.save(self.output_path)
        self.log.emit(f"\n{'='*50}")
        self.log.emit(f"✅ 处理完成！")
        self.log.emit(f"   共处理 {total} 行数据")
        self.log.emit(f"   成功插入 {total_images} 张图片")
        if failed_downloads:
            self.log.emit(f"   下载失败 {failed_downloads} 张")
        self.log.emit(f"   输出文件: {self.output_path}")
        self.log.emit(f"{'='*50}")

        self.finished.emit(True, f"处理完成！共处理 {total} 行，插入 {total_images} 张图片")

    def _extract_urls(self, cell_value):
        if not cell_value:
            return []
        parts = re.split(r'[,，\n\r]+', str(cell_value))
        return [p.strip() for p in parts if p.strip().startswith('http')]

    def _download_image(self, url, row, img_idx):
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()

            # 从 URL 中提取原始文件名
            parsed = urlparse(url)
            path = unquote(parsed.path)
            orig_name = os.path.basename(path)
            if not orig_name or '.' not in orig_name:
                orig_name = f"img_r{row}_{img_idx}.jpg"

            # 保存到 excel_images 目录
            save_path = os.path.join(self.img_dir, orig_name)

            # 如果文件名冲突，加序号
            if os.path.exists(save_path):
                name, ext = os.path.splitext(orig_name)
                save_path = os.path.join(self.img_dir, f"{name}_{img_idx}{ext}")

            with open(save_path, 'wb') as f:
                f.write(resp.content)

            # 返回临时路径（用于插入Excel）和实际文件名
            return save_path, os.path.basename(save_path)
        except Exception as e:
            self.log.emit(f"  ⚠ 下载失败: {e}")
            return None, None


# ============================================================
# 主窗口
# ============================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 图片处理器")
        self.setMinimumSize(700, 600)
        self.setStyleSheet(self._get_stylesheet())

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # ── 标题 ──
        title = QLabel("Excel 图片处理器")
        title.setObjectName("titleLabel")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # ── 文件设置 ──
        file_group = QGroupBox("文件设置")
        file_layout = QVBoxLayout(file_group)
        file_layout.setSpacing(10)

        # 输入文件
        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("输入Excel文件："))
        self.input_path_edit = QLineEdit()
        self.input_path_edit.setPlaceholderText("请选择 Excel 文件...")
        self.input_path_edit.setReadOnly(True)
        input_layout.addWidget(self.input_path_edit)
        self.browse_btn = QPushButton("浏览")
        self.browse_btn.setFixedWidth(80)
        self.browse_btn.clicked.connect(self._browse_file)
        input_layout.addWidget(self.browse_btn)
        file_layout.addLayout(input_layout)

        # 输出文件
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("输出Excel文件："))
        self.output_path_edit = QLineEdit()
        self.output_path_edit.setPlaceholderText("自动生成（默认文件名 + _年月日_时分秒）")
        self.output_path_edit.setReadOnly(True)
        output_layout.addWidget(self.output_path_edit)
        file_layout.addLayout(output_layout)

        layout.addWidget(file_group)

        # ── 图片链接设置 ──
        col_group = QGroupBox("图片链接设置")
        col_layout = QHBoxLayout(col_group)
        col_layout.addWidget(QLabel("图片链接所在列："))
        self.col_edit = QLineEdit()
        self.col_edit.setPlaceholderText("E")
        self.col_edit.setMaxLength(3)
        self.col_edit.setFixedWidth(80)
        self.col_edit.setText("E")
        col_layout.addWidget(self.col_edit)
        col_layout.addStretch()
        layout.addWidget(col_group)

        # ── 进度条 ──
        progress_group = QGroupBox("处理进度")
        progress_layout = QVBoxLayout(progress_group)
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        progress_layout.addWidget(self.progress_bar)
        layout.addWidget(progress_group)

        # ── 处理日志 ──
        log_group = QGroupBox("处理日志")
        log_layout = QVBoxLayout(log_group)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 10))
        log_layout.addWidget(self.log_text)
        layout.addWidget(log_group)

        # ── 按钮行 ──
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.process_btn = QPushButton("开始处理")
        self.process_btn.setFixedSize(180, 40)
        self.process_btn.setObjectName("processBtn")
        self.process_btn.clicked.connect(self._start_process)
        btn_layout.addWidget(self.process_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 状态栏
        self.statusBar().showMessage("就绪")

    # ── 样式表 ──
    def _get_stylesheet(self):
        return """
        QMainWindow {
            background-color: #f5f5f5;
        }
        QGroupBox {
            font-size: 14px;
            font-weight: bold;
            color: #333;
            border: 1px solid #d0d0d0;
            border-radius: 8px;
            margin-top: 12px;
            padding: 15px 12px 12px 12px;
            background-color: #ffffff;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            subcontrol-position: top left;
            padding: 2px 10px;
            background-color: #ffffff;
        }
        QLabel {
            font-size: 13px;
            color: #333;
        }
        #titleLabel {
            font-size: 22px;
            font-weight: bold;
            color: #2c3e50;
            padding: 8px 0;
        }
        QLineEdit {
            font-size: 13px;
            padding: 6px 10px;
            border: 1px solid #ccc;
            border-radius: 4px;
            background-color: #fafafa;
        }
        QLineEdit:focus {
            border-color: #4a90d9;
            background-color: #ffffff;
        }
        QLineEdit:read-only {
            background-color: #f0f0f0;
            color: #555;
        }
        QPushButton {
            font-size: 13px;
            padding: 6px 16px;
            border: 1px solid #ccc;
            border-radius: 4px;
            background-color: #f0f0f0;
            color: #333;
        }
        QPushButton:hover {
            background-color: #e3e3e3;
            border-color: #aaa;
        }
        QPushButton:pressed {
            background-color: #d5d5d5;
        }
        #processBtn {
            font-size: 15px;
            font-weight: bold;
            color: #ffffff;
            background-color: #4a90d9;
            border: none;
            border-radius: 6px;
        }
        #processBtn:hover {
            background-color: #357abd;
        }
        #processBtn:pressed {
            background-color: #2a5f9e;
        }
        #processBtn:disabled {
            background-color: #b0c4de;
            color: #e0e0e0;
        }
        QProgressBar {
            font-size: 12px;
            text-align: center;
            border: 1px solid #ccc;
            border-radius: 4px;
            background-color: #e8e8e8;
            height: 24px;
        }
        QProgressBar::chunk {
            background-color: #4a90d9;
            border-radius: 3px;
        }
        QTextEdit {
            font-size: 12px;
            border: 1px solid #ccc;
            border-radius: 4px;
            background-color: #1e1e1e;
            color: #d4d4d4;
            padding: 6px;
        }
        QStatusBar {
            font-size: 12px;
            color: #666;
        }
        """

    # ── 浏览文件 ──
    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx *.xlsm);;所有文件 (*)"
        )
        if path:
            self.input_path_edit.setText(path)
            self._auto_generate_output(path)

    def _auto_generate_output(self, input_path):
        """自动生成输出文件名"""
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(input_path)
        output = f"{base}_处理后_{now}{ext}"
        self.output_path_edit.setText(output)

    # ── 开始处理 ──
    def _start_process(self):
        input_path = self.input_path_edit.text()
        if not input_path or not os.path.exists(input_path):
            QMessageBox.warning(self, "提示", "请先选择输入的 Excel 文件")
            return

        col = self.col_edit.text().strip().upper()
        if not col or not re.match(r'^[A-Z]{1,3}$', col):
            QMessageBox.warning(self, "提示", "请输入正确的列字母（如 E、F、G）")
            return

        # 输出路径
        if not self.output_path_edit.text():
            self._auto_generate_output(input_path)
        output_path = self.output_path_edit.text()

        # 禁用按钮，清空日志
        self.process_btn.setEnabled(False)
        self.process_btn.setText("处理中...")
        self.log_text.clear()
        self.progress_bar.setValue(0)
        self.statusBar().showMessage("正在处理...")

        # 启动线程
        self.worker = WorkerThread(input_path, output_path, col, self)
        self.worker.progress.connect(self._update_progress)
        self.worker.log.connect(self._append_log)
        self.worker.finished.connect(self._on_finished)
        self.worker.start()

    def _update_progress(self, current, total):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.progress_bar.setFormat(f"{current}/{total}")

    def _append_log(self, msg):
        self.log_text.append(msg)
        # 自动滚动到底部
        cursor = self.log_text.textCursor()
        cursor.movePosition(QTextCursor.End)
        self.log_text.setTextCursor(cursor)

    def _on_finished(self, success, msg):
        self.process_btn.setEnabled(True)
        self.process_btn.setText("开始处理")
        self.statusBar().showMessage(msg)

        if success:
            QMessageBox.information(self, "完成", msg)
        else:
            QMessageBox.critical(self, "错误", msg)


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())