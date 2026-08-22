"""
将 Excel 中 E 列的图片链接转换为图片，并排放到对应单元格上
支持：单张图片、多张图片（逗号/换行分隔）、自动适配高度
"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
import requests
import os
import tempfile
import re

INPUT_FILE = "行政检查.xlsx"
OUTPUT_FILE = "行政检查.xlsx"
SHEET_NAME = "行政检查"
COL_INDEX = 5  # E 列
ROW_HEIGHT = 100
IMG_HEIGHT = 80  # 图片高度（px），留边距
IMG_MARGIN = 5   # 图片之间的间距（px）
# 1px ≈ 9525 EMU (96 DPI下)
EMU_PER_PX = 9525
TIMEOUT = 30

# 支持的图片扩展名
SUPPORTED_EXTS = {'jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'}

def extract_urls(cell_value):
    """从单元格文本中提取所有图片URL"""
    if not cell_value:
        return []
    # 统一按逗号、换行、分号分割
    parts = re.split(r'[,，\n\r]+', str(cell_value))
    return [p.strip() for p in parts if p.strip().startswith('http')]

def download_image(url, row, idx):
    """下载图片并返回临时文件路径"""
    try:
        resp = requests.get(url, timeout=TIMEOUT)
        resp.raise_for_status()
        # 提取扩展名
        match = re.search(r'\.([a-zA-Z]+)(?:\?|$)', url)
        ext = match.group(1).lower() if match else 'jpg'
        if ext not in SUPPORTED_EXTS:
            ext = 'jpg'
        tmp = os.path.join(tempfile.gettempdir(), f'xls_img_r{row}_i{idx}.{ext}')
        with open(tmp, 'wb') as f:
            f.write(resp.content)
        return tmp
    except Exception as e:
        print(f"  ⚠ 下载失败 (row {row}, url {idx}): {e}")
        return None

def main():
    print(f"📂 加载文件: {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    # 先统计所有有数据的行
    max_row = ws.max_row
    if max_row < 2:
        print("⚠ 没有数据行")
        return

    # 设置所有数据行的行高
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT

    # 设置 E 列宽度（根据图片数量自动调整）
    ws.column_dimensions['E'].width = 25

    total_images = 0
    failed_images = 0

    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=COL_INDEX)
        cell_value = cell.value
        urls = extract_urls(cell_value)

        if not urls:
            continue

        print(f"\n📌 第 {row} 行，发现 {len(urls)} 张图片链接")

        # 下载所有图片
        tmp_files = []
        for idx, url in enumerate(urls):
            print(f"  ⬇ 下载: {url}")
            tmp = download_image(url, row, idx)
            if tmp:
                tmp_files.append(tmp)
            else:
                failed_images += 1

        if not tmp_files:
            continue

        # 计算总宽度，调整列宽
        total_width_px = 0
        img_sizes = []
        for tmp in tmp_files:
            img = Image(tmp)
            scale = IMG_HEIGHT / img.height
            w = int(img.width * scale)
            img_sizes.append((w, IMG_HEIGHT))
            total_width_px += w + IMG_MARGIN
        total_width_px -= IMG_MARGIN  # 去掉最后一个间距

        # 调整列宽（Excel列宽单位 ≈ 7px 每单位）
        needed_col_width = max(15, total_width_px / 7 + 2)
        ws.column_dimensions['E'].width = needed_col_width

        # 逐张插入图片，并排排列
        for idx, tmp in enumerate(tmp_files):
            img = Image(tmp)
            w, h = img_sizes[idx]
            scale = IMG_HEIGHT / img.height
            img.width = int(img.width * scale)
            img.height = IMG_HEIGHT

            # 计算偏移量（EMU）
            # 之前的图片宽度 + 间距
            offset_x_emu = 0
            for i in range(idx):
                offset_x_emu += (img_sizes[i][0] + IMG_MARGIN) * EMU_PER_PX

            # 垂直居中微调
            offset_y_emu = int((ROW_HEIGHT - IMG_HEIGHT) / 2 * EMU_PER_PX)

            # 创建锚点：col=4 (E列, 0-based), row-2 (row 0-based)
            anchor = OneCellAnchor(
                _from=AnchorMarker(
                    col=COL_INDEX - 1,      # E列=4
                    row=row - 1,            # 0-based行号
                    colOff=offset_x_emu,
                    rowOff=offset_y_emu
                ),
                ext=XDRPositiveSize2D(
                    cx=int(img.width * EMU_PER_PX),
                    cy=int(img.height * EMU_PER_PX)
                )
            )
            img.anchor = anchor
            ws.add_image(img)
            total_images += 1
            print(f"  ✅ 插入图片 {idx + 1}: {img.width}x{img.height}px")

        # 图片插入后，把原链接文本保留在单元格中（作为备注）
        # 但为了不遮挡，可以把它放到批注里，或者保留在原位（图片会浮在文字上方）
        # 此处保留原值，图片会覆盖在文字上方

    # 保存
    wb.save(OUTPUT_FILE)
    print(f"\n{'='*50}")
    print(f"✅ 处理完成！")
    print(f"   共处理 {max_row - 1} 行数据")
    print(f"   成功插入 {total_images} 张图片")
    if failed_images:
        print(f"   失败 {failed_images} 张图片")
    print(f"   所有行高已设为 {ROW_HEIGHT}px")
    print(f"   输出文件: {OUTPUT_FILE}")
    print(f"{'='*50}")

if __name__ == '__main__':
    main()